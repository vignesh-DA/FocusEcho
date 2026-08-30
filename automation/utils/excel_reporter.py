"""
Excel Report Generator for FocusEcho Appium E2E Framework
Generates a multi-sheet .xlsx report using openpyxl.

Sheets:
  1. All Tests
  2. Passed Tests
  3. Failed Tests
  4. Skipped Tests
  5. Execution Metrics
  6. Defect Summary
  7. Pass Rate Summary
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠ openpyxl not installed. Run: pip install openpyxl")

from config.test_config import EXCEL_DIR, BUILD_NUMBER, GIT_COMMIT, BRANCH, DEVICE_NAME, ANDROID_VERSION, APP_VERSION


# ─────────────────────────────────────────────────────────────────────────────
# Colour Palette
# ─────────────────────────────────────────────────────────────────────────────
COLOR_HEADER_BG = "1E293B"   # Dark slate
COLOR_HEADER_FG = "FFFFFF"
COLOR_PASS = "16A34A"        # Green
COLOR_FAIL = "DC2626"        # Red
COLOR_SKIP = "D97706"        # Amber
COLOR_BLOCK = "6B7280"       # Gray
COLOR_ROW_ALT = "F1F5F9"     # Light blue-gray
COLOR_TITLE_BG = "0F172A"    # Very dark


def _make_fill(hex_color: str) -> "PatternFill":
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _header_font() -> "Font":
    return Font(bold=True, color=COLOR_HEADER_FG, size=11, name="Calibri")


def _title_font() -> "Font":
    return Font(bold=True, color=COLOR_HEADER_FG, size=14, name="Calibri")


def _thin_border() -> "Border":
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


def _status_fill(status: str) -> "PatternFill":
    mapping = {
        "PASS": COLOR_PASS,
        "PASSED": COLOR_PASS,
        "FAIL": COLOR_FAIL,
        "FAILED": COLOR_FAIL,
        "SKIP": COLOR_SKIP,
        "SKIPPED": COLOR_SKIP,
        "BLOCKED": COLOR_BLOCK,
    }
    return _make_fill(mapping.get(status.upper(), "FFFFFF"))


class ExcelReporter:
    """Generates comprehensive Excel test reports."""

    COLUMNS = [
        "Test ID", "Module", "Test Name", "Priority",
        "Status", "Execution Time (s)", "Error Message",
        "Screenshot Path", "Timestamp"
    ]

    def __init__(self, results: list[dict[str, Any]]):
        """
        Args:
            results: List of test result dicts. Each must have:
                - test_id, module, name, priority, status,
                  duration, error_message, screenshot_path
        """
        self.results = results
        self.generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.passed = [r for r in results if r["status"].upper() in ("PASS", "PASSED")]
        self.failed = [r for r in results if r["status"].upper() in ("FAIL", "FAILED")]
        self.skipped = [r for r in results if r["status"].upper() in ("SKIP", "SKIPPED")]
        self.blocked = [r for r in results if r["status"].upper() == "BLOCKED"]
        self.total = len(results)
        self.pass_rate = (len(self.passed) / self.total * 100) if self.total else 0

    def generate(self) -> Path:
        """Generate the main Automation_Test_Report.xlsx."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel report generation")

        wb = openpyxl.Workbook()

        self._write_all_tests_sheet(wb.active)
        wb.active.title = "All Tests"
        self._write_results_sheet(wb.create_sheet("Passed Tests"), self.passed)
        self._write_results_sheet(wb.create_sheet("Failed Tests"), self.failed)
        self._write_results_sheet(wb.create_sheet("Skipped Tests"), self.skipped)
        self._write_metrics_sheet(wb.create_sheet("Execution Metrics"))
        self._write_defect_sheet(wb.create_sheet("Defect Summary"))
        self._write_pass_rate_sheet(wb.create_sheet("Pass Rate Summary"))

        out_path = EXCEL_DIR / "Automation_Test_Report.xlsx"
        wb.save(str(out_path))
        print(f"✅ Excel report saved: {out_path}")
        return out_path

    def generate_split_reports(self) -> dict[str, Path]:
        """Generate separate passed/failed/summary xlsx files."""
        paths = {}

        # Passed
        wb_pass = openpyxl.Workbook()
        self._write_results_sheet(wb_pass.active, self.passed)
        wb_pass.active.title = "Passed Tests"
        p = EXCEL_DIR / "Passed_Test_Cases.xlsx"
        wb_pass.save(str(p))
        paths["passed"] = p

        # Failed
        wb_fail = openpyxl.Workbook()
        self._write_results_sheet(wb_fail.active, self.failed)
        wb_fail.active.title = "Failed Tests"
        p = EXCEL_DIR / "Failed_Test_Cases.xlsx"
        wb_fail.save(str(p))
        paths["failed"] = p

        # Summary
        wb_sum = openpyxl.Workbook()
        self._write_metrics_sheet(wb_sum.active)
        wb_sum.active.title = "Execution Summary"
        p = EXCEL_DIR / "Execution_Summary.xlsx"
        wb_sum.save(str(p))
        paths["summary"] = p

        return paths

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet Writers
    # ─────────────────────────────────────────────────────────────────────────

    def _write_all_tests_sheet(self, ws) -> None:
        """Sheet 1: All executed test cases."""
        self._write_title_row(ws, "FocusEcho AI — Full Automation Test Report", len(self.COLUMNS))
        self._write_info_rows(ws, len(self.COLUMNS))
        header_row = ws.max_row + 1
        self._write_header_row(ws, self.COLUMNS, header_row)

        for i, result in enumerate(self.results, start=header_row + 1):
            row_data = [
                result.get("test_id", ""),
                result.get("module", ""),
                result.get("name", ""),
                result.get("priority", ""),
                result.get("status", ""),
                round(result.get("duration", 0), 2),
                result.get("error_message", ""),
                result.get("screenshot_path", ""),
                result.get("timestamp", ""),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if col == 5:  # Status column
                    cell.fill = _status_fill(str(value))
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                elif i % 2 == 0:
                    cell.fill = _make_fill(COLOR_ROW_ALT)

        self._auto_fit_columns(ws)

    def _write_results_sheet(self, ws, results: list) -> None:
        """Reusable sheet writer for passed/failed/skipped."""
        self._write_header_row(ws, self.COLUMNS, 1)
        for i, result in enumerate(results, start=2):
            row_data = [
                result.get("test_id", ""),
                result.get("module", ""),
                result.get("name", ""),
                result.get("priority", ""),
                result.get("status", ""),
                round(result.get("duration", 0), 2),
                result.get("error_message", ""),
                result.get("screenshot_path", ""),
                result.get("timestamp", ""),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = _thin_border()
                if col == 5:
                    cell.fill = _status_fill(str(value))
                    cell.font = Font(bold=True, color="FFFFFF")
        self._auto_fit_columns(ws)

    def _write_metrics_sheet(self, ws) -> None:
        """Sheet 5: Execution Metrics."""
        metrics = [
            ("Metric", "Value"),
            ("Build Number", BUILD_NUMBER),
            ("Git Commit", GIT_COMMIT),
            ("Branch", BRANCH),
            ("Device", DEVICE_NAME),
            ("Android Version", ANDROID_VERSION),
            ("App Version", APP_VERSION),
            ("Report Generated", self.generated_at),
            ("", ""),
            ("Total Test Cases", self.total),
            ("Executed", self.total),
            ("Passed", len(self.passed)),
            ("Failed", len(self.failed)),
            ("Skipped", len(self.skipped)),
            ("Blocked", len(self.blocked)),
            ("Pass Rate (%)", f"{self.pass_rate:.1f}%"),
            ("Fail Rate (%)", f"{(100 - self.pass_rate):.1f}%"),
            ("", ""),
            ("Avg Duration (s)",
             f"{(sum(r.get('duration', 0) for r in self.results) / max(self.total, 1)):.2f}"),
            ("Total Duration (s)",
             f"{sum(r.get('duration', 0) for r in self.results):.2f}"),
        ]
        for i, (key, val) in enumerate(metrics, 1):
            ws.cell(row=i, column=1, value=key).font = Font(bold=True, name="Calibri")
            ws.cell(row=i, column=2, value=val)
        self._auto_fit_columns(ws)

    def _write_defect_sheet(self, ws) -> None:
        """Sheet 6: Defect Summary — failed tests with details."""
        cols = ["Test ID", "Module", "Test Name", "Priority", "Error Message", "Screenshot"]
        self._write_header_row(ws, cols, 1)
        for i, result in enumerate(self.failed, start=2):
            ws.cell(row=i, column=1, value=result.get("test_id", ""))
            ws.cell(row=i, column=2, value=result.get("module", ""))
            ws.cell(row=i, column=3, value=result.get("name", ""))
            ws.cell(row=i, column=4, value=result.get("priority", ""))
            ws.cell(row=i, column=5, value=result.get("error_message", ""))
            ws.cell(row=i, column=6, value=result.get("screenshot_path", ""))
        self._auto_fit_columns(ws)

    def _write_pass_rate_sheet(self, ws) -> None:
        """Sheet 7: Pass Rate by module."""
        modules: dict[str, dict] = {}
        for r in self.results:
            mod = r.get("module", "Unknown")
            if mod not in modules:
                modules[mod] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0}
            modules[mod]["total"] += 1
            status = r.get("status", "").upper()
            if status in ("PASS", "PASSED"):
                modules[mod]["passed"] += 1
            elif status in ("FAIL", "FAILED"):
                modules[mod]["failed"] += 1
            else:
                modules[mod]["skipped"] += 1

        cols = ["Module", "Total", "Passed", "Failed", "Skipped", "Pass Rate (%)"]
        self._write_header_row(ws, cols, 1)
        for i, (mod, data) in enumerate(sorted(modules.items()), start=2):
            rate = (data["passed"] / data["total"] * 100) if data["total"] else 0
            ws.cell(row=i, column=1, value=mod)
            ws.cell(row=i, column=2, value=data["total"])
            ws.cell(row=i, column=3, value=data["passed"])
            ws.cell(row=i, column=4, value=data["failed"])
            ws.cell(row=i, column=5, value=data["skipped"])
            ws.cell(row=i, column=6, value=f"{rate:.1f}%")
        self._auto_fit_columns(ws)

    # ─────────────────────────────────────────────────────────────────────────
    # Helpers
    # ─────────────────────────────────────────────────────────────────────────

    def _write_title_row(self, ws, title: str, col_span: int) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
        cell = ws.cell(row=1, column=1, value=title)
        cell.fill = _make_fill(COLOR_TITLE_BG)
        cell.font = _title_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

    def _write_info_rows(self, ws, col_span: int) -> None:
        info = [
            f"Generated: {self.generated_at}   |   "
            f"Build: {BUILD_NUMBER}   |   "
            f"Device: {DEVICE_NAME}   |   "
            f"Android: {ANDROID_VERSION}   |   "
            f"App: {APP_VERSION}   |   "
            f"Pass Rate: {self.pass_rate:.1f}%"
        ]
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
        cell = ws.cell(row=2, column=1, value=info[0])
        cell.fill = _make_fill("334155")
        cell.font = Font(color="94A3B8", size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_header_row(self, ws, columns: list, row: int) -> None:
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = _make_fill(COLOR_HEADER_BG)
            cell.font = _header_font()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()

    def _auto_fit_columns(self, ws, max_width: int = 60) -> None:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def generate_all_reports(results: list[dict]) -> dict[str, str]:
    """
    Generate all Excel reports from test results.
    Returns dict mapping report name to file path.
    """
    reporter = ExcelReporter(results)
    main = reporter.generate()
    splits = reporter.generate_split_reports()
    return {
        "main": str(main),
        "passed": str(splits["passed"]),
        "failed": str(splits["failed"]),
        "summary": str(splits["summary"]),
    }
